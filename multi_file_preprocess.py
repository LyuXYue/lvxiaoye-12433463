import pandas as pd
import re
from openpyxl import load_workbook


def parse_excel_sheet(raw_file):
    """从Excel单元格读取原始文本数据"""
    wb = load_workbook(raw_file)
    sheet = wb.active
    raw_text = " ".join(str(cell.value) for row in sheet.iter_rows() for cell in row)
    return raw_text


def parse_codon_data(raw_text, species_name):
    """解析密码子数据文本"""
    pattern = re.compile(
        r'([A-Z]{3})\s+([A-Z\*])\s+([0-9.]+)\s+([0-9.]+)\s+\(([0-9]+)\)'
    )
    matches = pattern.findall(raw_text)

    data = []
    for triplet, aa, fraction, freq, num in matches:
        data.append({
            'Triplet': triplet,
            'Amino Acid': aa,
            'Fraction': float(fraction),
            'Frequency': float(freq),
            'Number': int(num.replace(' ', '')),
            'Species': species_name
        })

    return pd.DataFrame(data)


# 处理三个物种
species_files = {
    "Human": "human_raw.xlsx",
    "Mouse": "mouse_raw.xlsx",
    "Yeast": "yeast_raw.xlsx"
}

all_data = []
for species, filepath in species_files.items():
    try:
        print(f"正在处理 {species} 数据...")
        raw_text = parse_excel_sheet(filepath)
        df = parse_codon_data(raw_text, species)
        all_data.append(df)
        print(f"✅ {species} 数据处理完成")
    except Exception as e:
        print(f"❌ {species} 处理失败: {e}")

# 合并并保存结果
if all_data:
    combined_df = pd.concat(all_data)
    combined_df.to_excel("combined_codon_data.xlsx", index=False)
    print("\n所有数据处理完成！保存为 combined_codon_data.xlsx")
else:
    print("⚠️ 无有效数据处理结果")